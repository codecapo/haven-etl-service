"""Camden stock mapper.

The Camden export is block-level capital-works data: Block Code, Block Address
("1-160 Southfleet (Cons)"), Estate, Units, plus cost columns. No UPRNs. We
locate the block sheet, then explode each block into individual property rows.

Because there are no UPRNs, these rows can't be loaded into `property` (UPRN PK)
until address→UPRN matching exists — the pipeline routes them to the "unmatched"
output.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .base import PropertyRow, explode_block

COUNCIL = "camden"

# Header cells we look for (lowercased) to identify the block sheet + columns.
_ALIASES = {
    "block_code": {"block code", "blockcode"},
    "block_address": {"block address", "address", "block name"},
    "estate": {"estate", "street / estate", "street/estate"},
    "units": {"units", "no of units", "number of units", "no. of units"},
    "postcode": {"postcode", "post code", "postcode(s) found", "postcode found", "postcodes"},
    "uprn": {"uprn"},
}


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _first_postcode(v) -> str | None:
    """A single postcode from cells that may list several ('NW5 1NA / NW5 1NB')."""
    if not v:
        return None
    first = re.split(r"[\/,;]| or ", str(v).strip(), maxsplit=1)[0].strip()
    return first or None


def _find_header(ws, max_scan: int = 15) -> tuple[int, dict[str, int]] | None:
    """Find the header row + a field→column-index map on a worksheet."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = {_norm(c): i for i, c in enumerate(row) if c is not None}
        mapping: dict[str, int] = {}
        for field, names in _ALIASES.items():
            for name, idx in cells.items():
                if name in names:
                    mapping[field] = idx
                    break
        # A real block sheet has at least a block address + (code or units).
        if "block_address" in mapping and ("block_code" in mapping or "units" in mapping):
            return r, mapping
    return None


def parse(path: str | Path) -> list[PropertyRow]:
    """Parse the Camden workbook → exploded individual property rows."""
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    for ws in wb.worksheets:
        found = _find_header(ws)
        if not found:
            continue
        header_row, col = found
        rows: list[PropertyRow] = []

        def cell(values, key):
            i = col.get(key)
            return values[i] if i is not None and i < len(values) else None

        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if values is None or all(v is None for v in values):
                continue
            block_address = cell(values, "block_address")
            if not block_address:
                continue
            units_raw = cell(values, "units")
            try:
                units = int(float(units_raw)) if units_raw not in (None, "") else None
            except (TypeError, ValueError):
                units = None
            rows.extend(
                explode_block(
                    council=COUNCIL,
                    block_code=(str(cell(values, "block_code")).strip() if cell(values, "block_code") else None),
                    block_address=str(block_address).strip(),
                    units=units,
                    estate=(str(cell(values, "estate")).strip() if cell(values, "estate") else None),
                    postcode=_first_postcode(cell(values, "postcode")),
                )
            )
        wb.close()
        return rows
    wb.close()
    raise ValueError(f"No recognizable block sheet found in {path}")
